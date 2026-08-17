"""Generate the minimal spreadsheet fixtures the reader tests run against.

The institutional readers parse layouts that vary between sections, so they
need real files to be tested at all. Rather than ship copies of student
records, these fixtures are written from scratch: three invented students,
the same block structure as the real templates, and no real data.

    python tests/make_fixtures.py
"""
from __future__ import annotations

import os

from openpyxl import Workbook

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "fixtures")

STUDENTS = [("900001", "甲同学"), ("900002", "乙同学"), ("900003", "丙同学")]


def gradebook(path: str) -> None:
    """Coursework register: 3 classroom sessions, 2 assignments, two sections
    with different block widths, so the width-discovery logic is exercised."""
    wb = Workbook()
    for si, (title, n_cls, n_hw) in enumerate(
            [("班级一", 3, 2), ("班级二", 2, 3)]):
        ws = wb.active if si == 0 else wb.create_sheet()
        ws.title = title
        ws.cell(1, 1, "学生平时成绩登记表")
        ws.cell(2, 1, "开课学期：2099-2100-1")
        ws.cell(2, 4, "课程：测试课程")
        ws.cell(2, 15, f"班级：{title}")
        ws.cell(3, 1, "序号"); ws.cell(3, 2, "学号"); ws.cell(3, 3, "姓名")
        c_start = 4
        ws.cell(3, c_start, "课堂表现")
        c_mean = c_start + n_cls
        h_start = c_mean + 1
        ws.cell(3, h_start, "平时作业")
        h_mean = h_start + n_hw
        t_col = h_mean + 1
        ws.cell(3, t_col, "平时总评")
        ws.cell(4, c_start, "比例20％"); ws.cell(4, h_start, "比例80％")
        for k in range(n_cls):
            ws.cell(5, c_start + k, "3月"); ws.cell(6, c_start + k, f"{k + 1}日")
        ws.cell(5, c_mean, "平均")
        for k in range(n_hw):
            ws.cell(5, h_start + k, "4月"); ws.cell(6, h_start + k, f"{k + 1}日")
        ws.cell(5, h_mean, "平均")
        for i, (sid, name) in enumerate(STUDENTS):
            r = 7 + i
            ws.cell(r, 1, i + 1); ws.cell(r, 2, sid); ws.cell(r, 3, name)
            cls = [60 + 10 * i + 5 * k for k in range(n_cls)]
            hw = [70 + 5 * i + 3 * k for k in range(n_hw)]
            for k, v in enumerate(cls):
                ws.cell(r, c_start + k, v)
            ws.cell(r, c_mean, sum(cls) / n_cls)
            for k, v in enumerate(hw):
                ws.cell(r, h_start + k, v)
            ws.cell(r, h_mean, sum(hw) / n_hw)
            ws.cell(r, t_col, 0.2 * sum(cls) / n_cls + 0.8 * sum(hw) / n_hw)
            ws.cell(r, t_col + 1, sid); ws.cell(r, t_col + 2, name)
            ws.cell(r, t_col + 3, 75 + 4 * i)          # final mark
    wb.save(path)


def transcript(path: str) -> None:
    """Registrar export: the block of columns repeats twice across the page,
    so two students share a printed row and the third starts a new one."""
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1, "学生成绩单")
    ws.cell(2, 1, "开课学期：2099-2100-1")
    ws.cell(3, 1, "课程号：T0001"); ws.cell(3, 5, "课程名称：测试课程")
    head = ["序号", "学号", "姓名", "班级", "平时成绩", "期中成绩",
            "期末成绩", "总评成绩", "备注"]
    for block in (0, 1):
        for j, h in enumerate(head):
            ws.cell(4, 1 + block * len(head) + j, h)
    for i, (sid, name) in enumerate(STUDENTS):
        r, block = 5 + i // 2, i % 2
        off = block * len(head)
        # Must agree with the coursework register for section 一, otherwise the
        # merge would legitimately report a conflict: daily = 0.2*classroom
        # mean + 0.8*assignment mean = 70.2 + 6i, final = 75 + 4i.
        daily = 70.2 + 6 * i
        final = 75 + 4 * i
        overall = float(int(0.5 * daily + 0.5 * final + 0.5))   # registrar: half-up
        vals = [i + 1, sid, name, "班级一", daily, None, final, overall, None]
        for j, v in enumerate(vals):
            if v is not None:
                ws.cell(r, 1 + off + j, v)
    wb.save(path)


if __name__ == "__main__":
    os.makedirs(OUT, exist_ok=True)
    gradebook(os.path.join(OUT, "gradebook-sample.xlsx"))
    transcript(os.path.join(OUT, "transcript-sample.xlsx"))
    print("fixtures written to", OUT)
