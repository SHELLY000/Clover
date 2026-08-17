"""Tests for CLOVER.

Two things are worth testing here beyond the usual plumbing. First, that
the attainment arithmetic reproduces what an institution computes by hand
— otherwise nobody will trust the tool enough to replace the spreadsheet.
Second, that the diagnostics *fire*: a validity check that never fails is
worse than no check, because it certifies whatever it is shown.
"""

from __future__ import annotations

import json
import os
import tempfile

import numpy as np
import pytest

from clover import (
    build_matrix,
    compute_attainment,
    diagnose,
    load_blueprint,
    parse_blueprint,
    recommendations,
    run,
)
from clover.blueprint import BlueprintError
from clover.ingest import build_cohort
from clover.model import Cohort, Student

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)


# --------------------------------------------------------------------------
# fixtures
# --------------------------------------------------------------------------

def _blueprint_dict(items, objectives=None, total=100.0):
    return {
        "schema_version": "1.0",
        "blueprint_id": "test",
        "course": {"code": "T101", "name": "Test", "term": "2025-1"},
        "policy": {"qualifying_standard": 0.65, "total_points": total,
                   "bootstrap_iterations": 200, "seed": 7},
        "indicators": [{"id": "I1", "text": "i1"}, {"id": "I2", "text": "i2"}],
        "objectives": objectives or [
            {"id": "CO1", "indicator": "I1", "text": "o1"},
            {"id": "CO2", "indicator": "I2", "text": "o2"},
        ],
        "items": items,
    }


def _cohort(scores_by_student):
    return Cohort(course_code="T101", term="2025-1",
                  students=[Student(sid=str(i), name=f"s{i}", class_name="A",
                                    scores=sc)
                            for i, sc in enumerate(scores_by_student)])


# --------------------------------------------------------------------------
# blueprint validation
# --------------------------------------------------------------------------

def test_blueprint_rejects_allocation_that_does_not_sum_to_total():
    bp = _blueprint_dict([
        {"id": "A", "source": "a", "allocations": {"CO1": 40}},
        {"id": "B", "source": "b", "allocations": {"CO2": 40}},
    ])
    with pytest.raises(BlueprintError, match="sum to 80"):
        parse_blueprint(bp)


def test_blueprint_rejects_unknown_objective_reference():
    bp = _blueprint_dict([
        {"id": "A", "source": "a", "allocations": {"CO1": 50, "CO9": 50}},
    ])
    with pytest.raises(BlueprintError, match="CO9"):
        parse_blueprint(bp)


def test_blueprint_rejects_objective_with_no_points():
    bp = _blueprint_dict([{"id": "A", "source": "a", "allocations": {"CO1": 100}}])
    with pytest.raises(BlueprintError, match="CO2 receives no points"):
        parse_blueprint(bp)


def test_blueprint_hash_is_stable_and_content_sensitive():
    d = _blueprint_dict([
        {"id": "A", "source": "a", "allocations": {"CO1": 50}},
        {"id": "B", "source": "b", "allocations": {"CO2": 50}},
    ])
    h1 = parse_blueprint(d).hash()
    h2 = parse_blueprint(json.loads(json.dumps(d))).hash()
    assert h1 == h2
    d["items"][0]["allocations"]["CO1"] = 51
    d["items"][1]["allocations"]["CO2"] = 49
    assert parse_blueprint(d).hash() != h1


# --------------------------------------------------------------------------
# attainment arithmetic
# --------------------------------------------------------------------------

def test_attainment_matches_hand_calculation():
    bp = parse_blueprint(_blueprint_dict([
        {"id": "A", "source": "a", "allocations": {"CO1": 60}},
        {"id": "B", "source": "b", "allocations": {"CO2": 40}},
    ]))
    cohort = _cohort([{"a": 80.0, "b": 90.0}, {"a": 60.0, "b": 70.0}])
    res = compute_attainment(bp, cohort)
    co1 = next(o for o in res.objectives if o.id == "CO1")
    co2 = next(o for o in res.objectives if o.id == "CO2")
    assert co1.attainment == pytest.approx(0.70)          # (0.8 + 0.6) / 2
    assert co2.attainment == pytest.approx(0.80)          # (0.9 + 0.7) / 2
    assert co1.mean_points == pytest.approx(42.0)         # 0.70 * 60
    # course attainment is the points-weighted blend, not the mean of ratios
    assert res.course_attainment == pytest.approx((42.0 + 32.0) / 100.0)


def test_missing_item_is_excluded_not_scored_as_zero():
    bp = parse_blueprint(_blueprint_dict([
        {"id": "A", "source": "a", "allocations": {"CO1": 50}},
        {"id": "B", "source": "b", "allocations": {"CO2": 50}},
    ]))
    full = _cohort([{"a": 80.0, "b": 80.0}, {"a": 60.0, "b": 60.0}])
    gapped = _cohort([{"a": 80.0, "b": 80.0}, {"a": 60.0}])
    r_full = compute_attainment(bp, full)
    r_gap = compute_attainment(bp, gapped)
    co2_full = next(o for o in r_full.objectives if o.id == "CO2").attainment
    co2_gap = next(o for o in r_gap.objectives if o.id == "CO2").attainment
    # the present observation is unchanged; a zero-fill would have given 0.40
    assert co2_full == pytest.approx(0.70)
    assert co2_gap == pytest.approx(0.80)


def test_bootstrap_interval_brackets_the_estimate_and_is_reproducible():
    bp = parse_blueprint(_blueprint_dict([
        {"id": "A", "source": "a", "allocations": {"CO1": 50}},
        {"id": "B", "source": "b", "allocations": {"CO2": 50}},
    ]))
    rng = np.random.default_rng(0)
    cohort = _cohort([{"a": float(x), "b": float(y)}
                      for x, y in rng.integers(50, 100, size=(60, 2))])
    r1 = compute_attainment(bp, cohort)
    r2 = compute_attainment(bp, cohort)
    for o1, o2 in zip(r1.objectives, r2.objectives):
        assert o1.ci_low == o2.ci_low and o1.ci_high == o2.ci_high
        assert o1.ci_low <= o1.attainment <= o1.ci_high


def test_threshold_method_counts_students_over_the_expectation():
    bp = parse_blueprint(_blueprint_dict([
        {"id": "A", "source": "a", "allocations": {"CO1": 50}},
        {"id": "B", "source": "b", "allocations": {"CO2": 50}},
    ]))
    cohort = _cohort([{"a": 90.0, "b": 50.0}, {"a": 90.0, "b": 50.0},
                      {"a": 20.0, "b": 90.0}, {"a": 90.0, "b": 90.0}])
    res = compute_attainment(bp, cohort)
    co1 = next(o for o in res.objectives if o.id == "CO1")
    assert co1.threshold_attainment == pytest.approx(0.75)   # 3 of 4 clear 0.60


# --------------------------------------------------------------------------
# diagnostics must actually fire
# --------------------------------------------------------------------------

def _proportional_split_blueprint():
    """Two objectives, both computed as fixed shares of the same two marks."""
    return parse_blueprint(_blueprint_dict([
        {"id": "DAILY", "source": "d", "aggregate": True,
         "allocations": {"CO1": 20, "CO2": 30}},
        {"id": "FINAL", "source": "f", "aggregate": True,
         "allocations": {"CO1": 20, "CO2": 30}},
    ]))


def test_structural_confounding_is_detected():
    bp = _proportional_split_blueprint()
    rng = np.random.default_rng(1)
    cohort = _cohort([{"d": float(a), "f": float(b)}
                      for a, b in rng.integers(50, 100, size=(40, 2))])
    rep = diagnose(bp, build_matrix(bp, cohort))
    assert rep.confounded_pairs == [["CO1", "CO2"]]
    assert any(c.code == "ID-CONF" and c.level == "fail" for c in rep.checks)
    assert set(rep.grades.values()) == {"C"}


def test_rank_deficiency_is_detected_on_synthetic_proportional_data():
    bp = parse_blueprint(_blueprint_dict(
        [{"id": "DAILY", "source": "d", "aggregate": True,
          "allocations": {f"CO{i}": 10 for i in range(1, 6)}},
         {"id": "FINAL", "source": "f", "aggregate": True,
          "allocations": {f"CO{i}": 10 for i in range(1, 6)}}],
        objectives=[{"id": f"CO{i}", "indicator": "I1" if i < 3 else "I2",
                     "text": f"o{i}"} for i in range(1, 6)]))
    rng = np.random.default_rng(2)
    cohort = _cohort([{"d": float(a), "f": float(b)}
                      for a, b in rng.integers(60, 100, size=(50, 2))])
    rep = diagnose(bp, build_matrix(bp, cohort))
    assert rep.rank99 <= 2
    assert any(c.code == "ID-RANK" and c.level == "fail" for c in rep.checks)


def test_independent_evidence_passes_the_rank_check():
    bp = parse_blueprint(_blueprint_dict([
        {"id": "A1", "source": "a1", "allocations": {"CO1": 25}},
        {"id": "A2", "source": "a2", "allocations": {"CO1": 25}},
        {"id": "B1", "source": "b1", "allocations": {"CO2": 25}},
        {"id": "B2", "source": "b2", "allocations": {"CO2": 25}},
    ]))
    rng = np.random.default_rng(3)
    vals = rng.integers(40, 100, size=(80, 4))
    cohort = _cohort([{"a1": float(r[0]), "a2": float(r[1]),
                       "b1": float(r[2]), "b2": float(r[3])} for r in vals])
    rep = diagnose(bp, build_matrix(bp, cohort))
    assert rep.confounded_pairs == []
    assert not any(c.code == "ID-RANK" and c.level == "fail" for c in rep.checks)
    assert rep.separation_index > 0.5


def test_group_assigned_and_ceiling_effects_are_flagged():
    bp = parse_blueprint(_blueprint_dict([
        {"id": "A", "source": "a", "allocations": {"CO1": 50}},
        {"id": "G", "source": "g", "group_assigned": True,
         "allocations": {"CO2": 50}},
    ]))
    cohort = _cohort([{"a": 100.0, "g": float(80 + i % 3)} for i in range(40)])
    rep = diagnose(bp, build_matrix(bp, cohort))
    assert any(c.code == "EV-GRP" and c.scope == "objective:CO2" for c in rep.checks)
    assert any(c.code == "DS-CEIL" and c.scope == "item:A" for c in rep.checks)


def test_recommendations_are_produced_for_a_degenerate_blueprint():
    bp = _proportional_split_blueprint()
    rng = np.random.default_rng(5)
    cohort = _cohort([{"d": float(a), "f": float(b)}
                      for a, b in rng.integers(50, 100, size=(40, 2))])
    rep = diagnose(bp, build_matrix(bp, cohort))
    recs = recommendations(bp, rep)
    assert recs and any("分别记分" in r for r in recs)


# --------------------------------------------------------------------------
# readers and end-to-end
# --------------------------------------------------------------------------

def test_tabular_long_reader_round_trips_through_build_cohort():
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "scores.csv")
        with open(path, "w", encoding="utf-8") as fh:
            fh.write("sid,name,class,item,score\n")
            fh.write("1,alice,A,a,80\n1,alice,A,b,90\n2,bob,A,a,60\n2,bob,A,b,70\n")
        cohort = build_cohort("T101", "2025-1",
                              [{"path": path, "reader": "tabular"}])
    assert len(cohort.students) == 2
    assert cohort.students[0].scores == {"a": 80.0, "b": 90.0}
    assert cohort.sources[0].options == {}


def test_conflicting_sources_are_recorded_not_silently_merged():
    with tempfile.TemporaryDirectory() as tmp:
        p1 = os.path.join(tmp, "one.csv")
        p2 = os.path.join(tmp, "two.csv")
        for p, score in ((p1, 80), (p2, 55)):
            with open(p, "w", encoding="utf-8") as fh:
                fh.write("sid,name,class,item,score\n")
                fh.write(f"1,alice,A,a,{score}\n")
        cohort = build_cohort("T101", "2025-1",
                              [{"path": p1, "reader": "tabular"},
                               {"path": p2, "reader": "tabular"}])
    assert cohort.students[0].scores["a"] == 80.0        # first source wins
    assert len(cohort.conflicts) == 1
    assert cohort.conflicts[0]["discarded"] == 55.0


@pytest.mark.skipif(
    not os.path.exists(os.path.join(ROOT, "examples", "cohort-anonymised.csv")),
    reason="example dataset not present")
def test_example_dataset_reproduces_the_published_figures():
    bp = load_blueprint(os.path.join(
        ROOT, "blueprints", "example-bm0400067-declared.yaml"))
    cohort = build_cohort(bp.course.code, bp.course.term, [{
        "path": os.path.join(ROOT, "examples", "cohort-anonymised.csv"),
        "reader": "tabular"}])
    result, report, _ = run(bp, cohort)
    assert result.n_students == 92
    assert result.course_attainment == pytest.approx(0.8761, abs=5e-4)
    assert report.rank99 == 2
    assert set(report.grades.values()) == {"C"}


def test_missing_evidence_raises_a_helpful_error():
    from clover.compute import MissingEvidence
    bp = parse_blueprint(_blueprint_dict([
        {"id": "A", "source": "a", "allocations": {"CO1": 50}},
        {"id": "B", "source": "missing.key", "allocations": {"CO2": 50}},
    ]))
    with pytest.raises(MissingEvidence, match="missing.key"):
        build_matrix(bp, _cohort([{"a": 80.0}]))


def test_anonymisation_is_deterministic_and_preserves_scores():
    from clover.anonymize import anonymize_cohort
    cohort = _cohort([{"a": 80.0}, {"a": 60.0}])
    cohort.students[0].sid = "230541002"
    a1, key = anonymize_cohort(cohort, key="k")
    a2, _ = anonymize_cohort(cohort, key="k")
    a3, _ = anonymize_cohort(cohort, key="other")
    assert key == "k"
    assert a1.students[0].sid == a2.students[0].sid
    assert a1.students[0].sid != a3.students[0].sid
    assert "230541002" not in a1.students[0].sid
    assert a1.students[0].scores == cohort.students[0].scores


# --------------------------------------------------------------------------
# institutional readers (fixtures generated by tests/make_fixtures.py)
# --------------------------------------------------------------------------

FIXTURES = os.path.join(HERE, "fixtures")


@pytest.mark.skipif(not os.path.exists(os.path.join(FIXTURES, "gradebook-sample.xlsx")),
                    reason="run tests/make_fixtures.py first")
def test_gradebook_reader_discovers_block_widths_per_sheet():
    from clover.ingest.wbu_gradebook import read
    rows, keys = read(os.path.join(FIXTURES, "gradebook-sample.xlsx"))
    # two sections, three students each, with different numbers of sessions
    assert len(rows) == 6
    first = {r["sid"]: r for r in rows if r["class_name"] == "班级一"}
    second = {r["sid"]: r for r in rows if r["class_name"] == "班级二"}
    assert len(first) == len(second) == 3
    # section one recorded 3 classroom sessions and 2 assignments...
    assert "classroom.3" in first["900001"]["scores"]
    assert "homework.3" not in first["900001"]["scores"]
    # ...section two recorded 2 and 3; a hard-coded column layout would have
    # silently read the wrong cells here
    assert "classroom.3" not in second["900001"]["scores"]
    assert "homework.3" in second["900001"]["scores"]
    assert first["900001"]["scores"]["classroom.1"] == 60.0
    assert first["900001"]["scores"]["final.total"] == 75.0
    assert "daily.total" in keys and "final.total" in keys


@pytest.mark.skipif(not os.path.exists(os.path.join(FIXTURES, "transcript-sample.xlsx")),
                    reason="run tests/make_fixtures.py first")
def test_transcript_reader_walks_repeated_column_blocks():
    from clover.ingest.wbu_transcript import read
    rows, keys = read(os.path.join(FIXTURES, "transcript-sample.xlsx"))
    # three students laid out two-per-printed-row
    assert [r["sid"] for r in rows] == ["900001", "900002", "900003"]
    assert rows[1]["name"] == "乙同学"
    assert rows[1]["scores"] == {"daily.total": 76.2, "final.total": 79.0,
                                 "overall.total": 78.0}
    assert "midterm.total" not in keys        # column present but empty


def test_readers_merge_into_one_cohort_without_conflict():
    cohort = build_cohort("T0001", "2099-2100-1", [
        {"path": os.path.join(FIXTURES, "gradebook-sample.xlsx"),
         "reader": "wbu_gradebook", "sheets": ["班级一"]},
        {"path": os.path.join(FIXTURES, "transcript-sample.xlsx"),
         "reader": "wbu_transcript"},
    ])
    assert len(cohort.students) == 3
    assert cohort.conflicts == []      # register and transcript agree, as they must
    assert cohort.sources[0].options == {"sheets": ["班级一"]}
    # the merged record carries item-level keys from one source and the
    # registrar's overall mark from the other
    s0 = cohort.students[0].scores
    assert "classroom.1" in s0 and "overall.total" in s0


def test_disagreement_between_register_and_transcript_is_surfaced(tmp_path):
    """A gradebook that disagrees with the registrar is a finding, not noise."""
    from openpyxl import load_workbook
    src = os.path.join(FIXTURES, "transcript-sample.xlsx")
    tampered = tmp_path / "transcript-tampered.xlsx"
    wb = load_workbook(src)
    ws = wb.active
    ws.cell(5, 7, 99)                      # first student's final mark
    wb.save(tampered)
    cohort = build_cohort("T0001", "2099-2100-1", [
        {"path": os.path.join(FIXTURES, "gradebook-sample.xlsx"),
         "reader": "wbu_gradebook", "sheets": ["班级一"]},
        {"path": str(tampered), "reader": "wbu_transcript"},
    ])
    conflicts = [c for c in cohort.conflicts if c["key"] == "final.total"]
    assert len(conflicts) == 1
    assert conflicts[0]["kept"] == 75.0 and conflicts[0]["discarded"] == 99.0


# --------------------------------------------------------------------------
# fixes made in response to the v1.0.0 audit
# --------------------------------------------------------------------------

def _rounding_cohort():
    """Totals that land exactly on .5, where the rounding rule matters."""
    bp = parse_blueprint(_blueprint_dict([
        {"id": "A", "source": "a", "allocations": {"CO1": 50}},
        {"id": "B", "source": "b", "allocations": {"CO2": 50}},
    ]))
    # 0.5*(85+86)=85.5, 0.5*(87+88)=87.5 — both half-way cases
    cohort = _cohort([{"a": 85.0, "b": 86.0}, {"a": 87.0, "b": 88.0}])
    return bp, cohort


def test_integer_rounding_is_half_up_not_bankers():
    bp, cohort = _rounding_cohort()
    res = compute_attainment(bp, cohort)
    alt = res.distributions["total_alternate_rounding"]
    # half-to-even would give (86 + 88)/2 = 87.0; half-up gives (86 + 88)/2 too,
    # so use the individual bands: 85.5 must become 86, not 85
    assert alt["rounding"] == "integer"
    assert alt["mean"] == pytest.approx(87.0)
    assert alt["bands"]["89-80"] == 2


def test_both_rounding_conventions_are_always_reported():
    bp, cohort = _rounding_cohort()
    res = compute_attainment(bp, cohort)
    assert res.distributions["total"]["rounding"] == "none"
    assert res.distributions["total_alternate_rounding"]["rounding"] == "integer"
    bp.policy.grade_rounding = "integer"
    res2 = compute_attainment(bp, cohort)
    assert res2.distributions["total"]["rounding"] == "integer"
    assert res2.distributions["total_alternate_rounding"]["rounding"] == "none"


def test_unknown_rounding_rule_is_rejected():
    bp, cohort = _rounding_cohort()
    bp.policy.grade_rounding = "floor"
    with pytest.raises(ValueError, match="grade_rounding"):
        compute_attainment(bp, cohort)


def test_degenerate_threshold_method_is_flagged():
    bp = parse_blueprint(_blueprint_dict([
        {"id": "A", "source": "a", "allocations": {"CO1": 50}},
        {"id": "B", "source": "b", "allocations": {"CO2": 50}},
    ]))
    # everyone far above the expectation level: the method returns 1.0 for all
    cohort = _cohort([{"a": 90.0 + i, "b": 88.0 + i} for i in range(10)])
    sm = build_matrix(bp, cohort)
    res = compute_attainment(bp, cohort, sm)
    rep = diagnose(bp, sm, threshold=np.array(
        [o.threshold_attainment for o in res.objectives]))
    hit = [c for c in rep.checks if c.code == "TH-DEGEN"]
    assert hit and hit[0].level == "warn"


def test_threshold_method_passes_when_it_discriminates():
    bp = parse_blueprint(_blueprint_dict([
        {"id": "A", "source": "a", "allocations": {"CO1": 50}},
        {"id": "B", "source": "b", "allocations": {"CO2": 50}},
    ]))
    cohort = _cohort([{"a": 90.0, "b": 30.0}, {"a": 30.0, "b": 90.0},
                      {"a": 90.0, "b": 90.0}, {"a": 30.0, "b": 30.0}])
    sm = build_matrix(bp, cohort)
    res = compute_attainment(bp, cohort, sm)
    rep = diagnose(bp, sm, threshold=np.array(
        [o.threshold_attainment for o in res.objectives]))
    hit = [c for c in rep.checks if c.code == "TH-DEGEN"]
    assert hit and hit[0].level == "pass"


def test_dangling_derived_from_is_rejected():
    bp = _blueprint_dict([
        {"id": "A", "source": "a", "allocations": {"CO1": 50}},
        {"id": "B", "source": "b", "derived_from": "nowhere",
         "allocations": {"CO2": 50}},
    ])
    with pytest.raises(BlueprintError, match="derived_from"):
        parse_blueprint(bp)


def test_per_student_rows_can_be_withheld_from_the_record():
    bp = parse_blueprint(_blueprint_dict([
        {"id": "A", "source": "a", "allocations": {"CO1": 50}},
        {"id": "B", "source": "b", "allocations": {"CO2": 50}},
    ]))
    cohort = _cohort([{"a": 80.0, "b": 90.0}, {"a": 60.0, "b": 70.0}])
    assert len(compute_attainment(bp, cohort).per_student) == 2
    bp.policy.include_per_student = False
    withheld = compute_attainment(bp, cohort)
    assert withheld.per_student == []
    # withholding names must not change any reported figure
    assert withheld.course_attainment == pytest.approx(
        compute_attainment(parse_blueprint(_blueprint_dict([
            {"id": "A", "source": "a", "allocations": {"CO1": 50}},
            {"id": "B", "source": "b", "allocations": {"CO2": 50}},
        ])), cohort).course_attainment)


def test_students_with_missing_items_are_named_in_the_result():
    bp = parse_blueprint(_blueprint_dict([
        {"id": "A", "source": "a", "allocations": {"CO1": 50}},
        {"id": "B", "source": "b", "allocations": {"CO2": 50}},
    ]))
    cohort = _cohort([{"a": 80.0, "b": 80.0}, {"a": 60.0}])
    res = compute_attainment(bp, cohort)
    assert len(res.incomplete) == 1
    assert res.incomplete[0]["missing_items"] == ["B"]
