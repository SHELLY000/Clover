"""Reader for the registrar's transcript export (``学生成绩单``, legacy ``.xls``).

The export repeats the same block of columns twice across the page so
that two students fit on one printed row, so the reader locates *every*
occurrence of the ``学号`` header and walks each block independently.

Both the legacy binary ``.xls`` produced by the student-information system
and the ``.xlsx`` an instructor gets after re-saving it are accepted; the
sheet is read through a small uniform adapter so that the parsing logic
below sees the same interface either way. That is not only convenience —
it is what allows the reader to be covered by tests, since a fixture in
the modern format can be written from Python without a legacy encoder.

Emitted source keys: ``daily.total``, ``midterm.total``, ``final.total``,
``overall.total``.
"""

from __future__ import annotations

import re
from typing import Any

_COLS = {
    "学号": "sid",
    "姓名": "name",
    "班级": "class_name",
    "平时成绩": "daily.total",
    "期中成绩": "midterm.total",
    "期末成绩": "final.total",
    "总评成绩": "overall.total",
}


class _Sheet:
    """Uniform read-only view over an xlrd or an openpyxl worksheet."""

    __slots__ = ("_get", "ncols", "nrows")

    def __init__(self, nrows: int, ncols: int, get):
        self.nrows, self.ncols, self._get = nrows, ncols, get

    def cell_value(self, r: int, c: int) -> Any:
        """Zero-indexed, empty cells as the empty string."""
        v = self._get(r, c)
        return "" if v is None else v


def _sheets(path: str) -> list[_Sheet]:
    if path.lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=True)
        out = []
        for ws in wb.worksheets:
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            ncols = max((len(r) for r in rows), default=0)

            def get(r, c, rows=rows, ncols=ncols):
                if r >= len(rows) or c >= len(rows[r]):
                    return None
                return rows[r][c]

            out.append(_Sheet(len(rows), ncols, get))
        return out

    import xlrd
    wb = xlrd.open_workbook(path)
    return [_Sheet(sh.nrows, sh.ncols,
                   (lambda r, c, sh=sh: sh.cell_value(r, c)))
            for sh in wb.sheets()]


def _num(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _header_rows(sh) -> list[int]:
    out = []
    for r in range(min(sh.nrows, 20)):
        row = [str(sh.cell_value(r, c)).strip() for c in range(sh.ncols)]
        if "学号" in row and "姓名" in row:
            out.append(r)
    return out


def _blocks(sh, header_row: int) -> list[dict[str, int]]:
    """Split the header row into repeated column blocks keyed by field."""
    blocks: list[dict[str, int]] = []
    current: dict[str, int] = {}
    for c in range(sh.ncols):
        label = str(sh.cell_value(header_row, c)).strip()
        field = _COLS.get(label)
        if field is None:
            continue
        if field == "sid" and current:
            blocks.append(current)
            current = {}
        current[field] = c
    if current:
        blocks.append(current)
    return [b for b in blocks if "sid" in b]


def _course_meta(sh) -> dict[str, str]:
    meta: dict[str, str] = {}
    for r in range(min(sh.nrows, 6)):
        for c in range(sh.ncols):
            v = str(sh.cell_value(r, c) or "")
            for label, key in (("课程号", "code"), ("课程名称", "name"),
                               ("开课学期", "term"), ("学时", "hours")):
                m = re.search(rf"{label}[:：]\s*([^\s]+)", v)
                if m:
                    meta.setdefault(key, m.group(1).strip())
    return meta


def read(path: str, **_: Any) -> tuple[list[dict[str, Any]], list[str]]:
    rows: list[dict[str, Any]] = []
    keys: set[str] = set()

    for sh in _sheets(path):
        headers = _header_rows(sh)
        if not headers:
            continue
        header_row = headers[0]
        blocks = _blocks(sh, header_row)
        for r in range(header_row + 1, sh.nrows):
            for block in blocks:
                raw_sid = sh.cell_value(r, block["sid"])
                sid = str(raw_sid).strip()
                sid = sid.removesuffix(".0")
                if not sid.isdigit() or len(sid) < 6:
                    continue
                scores: dict[str, float] = {}
                for field, col in block.items():
                    if "." not in field:
                        continue
                    v = _num(sh.cell_value(r, col))
                    if v is not None:
                        scores[field] = v
                if not scores:
                    continue
                name_col = block.get("name")
                cls_col = block.get("class_name")
                rows.append({
                    "sid": sid,
                    "name": (str(sh.cell_value(r, name_col)).strip()
                             if name_col is not None else ""),
                    "class_name": (str(sh.cell_value(r, cls_col)).strip()
                                   if cls_col is not None else ""),
                    "scores": scores,
                })
                keys.update(scores)

    return rows, sorted(keys)
