"""Reader for the item-level coursework register (``平时成绩登记表``).

Layout produced by the institutional template, per worksheet:

===========  ==========================================================
columns      meaning
===========  ==========================================================
A..C         index, student id, student name
D..(D+p-1)   one column per classroom-participation session
next         session mean (recomputed, not trusted)
following q  one column per homework assignment
next         homework mean
next         coursework total
+2 .. +4     student id, name, final coursework mark
===========  ==========================================================

The block widths ``p`` and ``q`` vary between classes because different
sections met on different dates, so they are discovered from the header
rows rather than hard-coded.  Emitted source keys:

``classroom.1`` … ``classroom.p``, ``classroom.mean``,
``homework.1`` … ``homework.q``, ``homework.mean``,
``daily.total``, ``final.total``.
"""

from __future__ import annotations

import re
from typing import Any

from openpyxl import load_workbook

_HEADER_CLASSROOM = "课堂表现"
_HEADER_HOMEWORK = "平时作业"
_HEADER_TOTAL = "平时总评"
_MEAN = "平均"


def _find_header_row(ws) -> int:
    for r in range(1, min(ws.max_row, 12) + 1):
        for c in range(1, min(ws.max_column, 12) + 1):
            if str(ws.cell(r, c).value or "").strip() == "序号":
                return r
    raise ValueError("could not locate the header row (no 序号 cell found)")


def _label_columns(ws, header_row: int) -> dict[str, int]:
    found: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(header_row, c).value or "").strip()
        if v in (_HEADER_CLASSROOM, _HEADER_HOMEWORK, _HEADER_TOTAL) and v not in found:
            found[v] = c
    return found


def _mean_column(ws, start: int, limit: int, sub_row: int) -> int | None:
    for c in range(start, limit + 1):
        if str(ws.cell(sub_row, c).value or "").strip() == _MEAN:
            return c
    return None


def _class_name(ws) -> str:
    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(r, c).value or "")
            m = re.search(r"班级[:：]\s*(.+)", v)
            if m:
                return m.group(1).strip()
    return ws.title


def _is_student_id(v: Any) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    s = s.removesuffix(".0")
    return s.isdigit() and len(s) >= 6


def _num(v: Any) -> float | None:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def read(path: str, sheets: list[str] | None = None,
         skip_sheets: list[str] | None = None,
         **_: Any) -> tuple[list[dict[str, Any]], list[str]]:
    wb = load_workbook(path, data_only=True)
    names = sheets or [n for n in wb.sheetnames
                       if not skip_sheets or n not in skip_sheets]

    rows: list[dict[str, Any]] = []
    keys: set[str] = set()

    for name in names:
        ws = wb[name]
        header_row = _find_header_row(ws)
        labels = _label_columns(ws, header_row)
        if _HEADER_CLASSROOM not in labels or _HEADER_HOMEWORK not in labels:
            continue
        c_start = labels[_HEADER_CLASSROOM]
        h_start = labels[_HEADER_HOMEWORK]
        t_col = labels.get(_HEADER_TOTAL)

        sub_row = header_row + 2          # the row carrying 月 / 平均 markers
        c_mean = _mean_column(ws, c_start, h_start - 1, sub_row)
        h_limit = (t_col - 1) if t_col else ws.max_column
        h_mean = _mean_column(ws, h_start, h_limit, sub_row)
        c_last = (c_mean - 1) if c_mean else (h_start - 1)
        h_last = (h_mean - 1) if h_mean else h_limit

        n_class = c_last - c_start + 1
        n_home = h_last - h_start + 1
        cls = _class_name(ws)

        first_data = header_row + 4
        for r in range(first_data, ws.max_row + 1):
            sid_cell = ws.cell(r, 2).value
            if not _is_student_id(sid_cell):
                continue
            sid = str(sid_cell).strip()
            sid = sid.removesuffix(".0")
            scores: dict[str, float] = {}
            for k in range(n_class):
                v = _num(ws.cell(r, c_start + k).value)
                if v is not None:
                    scores[f"classroom.{k + 1}"] = v
            if c_mean:
                v = _num(ws.cell(r, c_mean).value)
                if v is not None:
                    scores["classroom.mean"] = v
            for k in range(n_home):
                v = _num(ws.cell(r, h_start + k).value)
                if v is not None:
                    scores[f"homework.{k + 1}"] = v
            if h_mean:
                v = _num(ws.cell(r, h_mean).value)
                if v is not None:
                    scores["homework.mean"] = v
            if t_col:
                v = _num(ws.cell(r, t_col).value)
                if v is not None:
                    scores["daily.total"] = v
                v = _num(ws.cell(r, t_col + 3).value)
                if v is not None:
                    scores["final.total"] = v
            rows.append({"sid": sid,
                         "name": str(ws.cell(r, 3).value or "").strip(),
                         "class_name": cls,
                         "scores": scores})
            keys.update(scores)

    return rows, sorted(keys)
