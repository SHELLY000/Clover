"""Writer for the institutional attainment calculation workbook.

Reproduces the four-header-row layout that reviewers expect: indicator
row, objective row, assessment-component row, target-points row, then one
row per student, then cohort mean, indicator attainment, objective
attainment and course attainment.  One worksheet per class plus a
combined sheet, exactly as the manual template is organised.

Two columns are added that the manual template does not have — the
bootstrap interval and the identifiability grade — because a number
without a stated precision and a number that cannot be measured look
identical on paper, and that is the whole problem this tool exists to fix.
"""

from __future__ import annotations

from typing import Any

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..blueprint import Blueprint
from ..compute import ScoreMatrix
from ..model import AttainmentResult, DiagnosticReport

_THIN = Side(style="thin", color="9AA0A6")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEAD_FILL = PatternFill("solid", fgColor="E8EEF7")
_NOTE_FILL = PatternFill("solid", fgColor="FFF4E5")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_GRADE_FILL = {"A": PatternFill("solid", fgColor="E3F3E6"),
               "B": PatternFill("solid", fgColor="FFF4E5"),
               "C": PatternFill("solid", fgColor="FBE3E3")}


def _columns(bp: Blueprint) -> list[tuple[str, str, str, float]]:
    """(objective_id, indicator_id, component label, target points) per column."""
    cols = []
    for o in bp.objectives:
        for it in bp.items_for_objective(o.id):
            cols.append((o.id, o.indicator, it.name, it.allocations[o.id]))
    return cols


def write_workbook(path: str, bp: Blueprint, sm: ScoreMatrix,
                   result: AttainmentResult,
                   report: DiagnosticReport | None = None,
                   by_class: bool = True) -> str:
    cols = _columns(bp)
    item_index = {it.id: i for i, it in enumerate(bp.items)}
    obj_index = {o: j for j, o in enumerate(sm.objective_ids)}

    wb = Workbook()
    wb.remove(wb.active)

    groups: list[tuple[str, list[int]]] = []
    if by_class:
        seen: dict[str, list[int]] = {}
        for s, st in enumerate(sm.students):
            seen.setdefault(st.class_name or "未分班", []).append(s)
        groups.extend(seen.items())
    groups.append(("合并", list(range(len(sm.students)))))

    for sheet_name, rows in groups:
        ws = wb.create_sheet(title=str(sheet_name)[:31])
        _write_sheet(ws, bp, sm, result, report, cols, item_index, obj_index, rows,
                     sheet_name)

    if report is not None:
        _write_diagnostics_sheet(wb.create_sheet("识别性诊断"), bp, result, report)

    wb.save(path)
    return path


def _write_sheet(ws, bp: Blueprint, sm: ScoreMatrix, result: AttainmentResult,
                 report: DiagnosticReport | None,
                 cols, item_index, obj_index, rows: list[int],
                 sheet_name: str) -> None:
    c = bp.course
    ncol = 3 + len(cols) + 4
    ws.cell(1, 1, f"{c.institution}课程目标、毕业要求达成度计算表").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws.cell(1, 1).alignment = _CENTER
    ws.cell(2, 1, f"专业：{c.program}    班级：{sheet_name}    "
                  f"课程：{c.name}（{c.code}）    学期：{c.term}    "
                  f"任课教师：{c.instructor}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)

    r_ind, r_obj, r_cmp, r_pts = 3, 4, 5, 6
    for r, label in ((r_ind, "毕业要求指标点"), (r_obj, "课程目标"),
                     (r_cmp, "评价内容"), (r_pts, "目标分值")):
        ws.cell(r, 3, label).font = Font(bold=True)
    ws.cell(r_ind, 1, "序号"); ws.cell(r_ind, 2, "学号")
    ws.merge_cells(start_row=r_ind, start_column=1, end_row=r_pts, end_column=1)
    ws.merge_cells(start_row=r_ind, start_column=2, end_row=r_pts, end_column=2)
    ws.cell(r_ind, 3, "学生姓名")
    ws.merge_cells(start_row=r_ind, start_column=3, end_row=r_pts, end_column=3)

    base = 4
    for i, (oid, ind, comp, pts) in enumerate(cols):
        col = base + i
        ws.cell(r_ind, col, f"指标点{ind}")
        ws.cell(r_obj, col, oid)
        ws.cell(r_cmp, col, comp)
        ws.cell(r_pts, col, round(pts, 4))
    for r in range(r_ind, r_pts + 1):
        for col in range(1, base + len(cols)):
            cell = ws.cell(r, col)
            cell.alignment = _CENTER
            cell.border = _BORDER
            cell.fill = _HEAD_FILL

    tail = base + len(cols)
    for k, label in enumerate(["总评成绩", "平时成绩", "期末成绩", "折合总评"]):
        cell = ws.cell(r_pts, tail + k, label)
        cell.font = Font(bold=True)
        cell.alignment = _CENTER
        cell.border = _BORDER
        cell.fill = _HEAD_FILL

    final_src = next((it.source for it in bp.items if it.kind == "final"), None)

    r = r_pts + 1
    for n, s in enumerate(rows, start=1):
        st = sm.students[s]
        ws.cell(r, 1, n); ws.cell(r, 2, st.sid); ws.cell(r, 3, st.name)
        for i, (oid, _ind, comp, _pts) in enumerate(cols):
            item = next(it for it in bp.items_for_objective(oid) if it.name == comp)
            val = sm.R[s, item_index[item.id]] * item.allocations[oid]
            ws.cell(r, base + i, round(float(val), 2))
        total = float(sm.P[s].sum())
        ws.cell(r, tail, round(total, 2))
        ws.cell(r, tail + 1, st.scores.get("daily.total"))
        ws.cell(r, tail + 2, st.scores.get(final_src) if final_src else None)
        ws.cell(r, tail + 3, st.scores.get("overall.total"))
        for col in range(1, tail + 4):
            ws.cell(r, col).border = _BORDER
            ws.cell(r, col).alignment = _CENTER
        r += 1

    idx = np.array(rows, dtype=int)
    r_mean, r_ind_att, r_obj_att, r_ci, r_grade, r_course = r, r + 1, r + 2, r + 3, r + 4, r + 5
    labels = {r_mean: "平均分", r_ind_att: "毕业要求指标点达成度",
              r_obj_att: "课程目标达成度", r_ci: "95%自助置信区间",
              r_grade: "识别性等级", r_course: "课程达成度"}
    for rr, label in labels.items():
        ws.cell(rr, 1, label).font = Font(bold=True)
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=3)
        ws.cell(rr, 1).alignment = _CENTER

    for i, (oid, ind, comp, pts) in enumerate(cols):
        item = next(it for it in bp.items_for_objective(oid) if it.name == comp)
        vals = sm.R[idx, item_index[item.id]] * pts
        ws.cell(r_mean, base + i, round(float(np.nanmean(vals)), 3))

    # objective-level rows are written across the span of each objective
    span_start = base
    for oid in sm.objective_ids:
        n_cols = sum(1 for cc in cols if cc[0] == oid)
        if n_cols == 0:
            continue
        j = obj_index[oid]
        sub = sm.ratio[idx, j]
        att = float(np.nanmean(sub))
        ind_id = bp.objective(oid).indicator
        ind_cols = [jj for jj, o2 in enumerate(sm.objective_ids)
                    if bp.objective(o2).indicator == ind_id]
        ind_tgt = sm.target[ind_cols].sum()
        ind_att = float(np.nanmean(sm.P[np.ix_(idx, ind_cols)].sum(axis=1)) / ind_tgt)

        for rr, val in ((r_ind_att, round(ind_att, 4)), (r_obj_att, round(att, 4))):
            ws.cell(rr, span_start, val)
            if n_cols > 1:
                ws.merge_cells(start_row=rr, start_column=span_start,
                               end_row=rr, end_column=span_start + n_cols - 1)
            ws.cell(rr, span_start).alignment = _CENTER

        obj_res = next(o for o in result.objectives if o.id == oid)
        ci = f"[{obj_res.ci_low:.3f}, {obj_res.ci_high:.3f}]" if sheet_name == "合并" else "—"
        ws.cell(r_ci, span_start, ci)
        grade = (report.grades.get(oid, "?") if report else "?")
        gcell = ws.cell(r_grade, span_start, grade)
        gcell.fill = _GRADE_FILL.get(grade, _HEAD_FILL)
        gcell.font = Font(bold=True)
        for rr in (r_ci, r_grade):
            if n_cols > 1:
                ws.merge_cells(start_row=rr, start_column=span_start,
                               end_row=rr, end_column=span_start + n_cols - 1)
            ws.cell(rr, span_start).alignment = _CENTER
        span_start += n_cols

    course_att = float(np.nanmean(sm.P[idx].sum(axis=1)) / sm.target.sum())
    ws.cell(r_course, base, round(course_att, 4)).font = Font(bold=True)
    ws.cell(r_course, base + 1,
            f"合格标准 {bp.policy.qualifying_standard:.2f}｜"
            f"{'达成' if course_att >= bp.policy.qualifying_standard else '未达成'}")

    for rr in labels:
        for col in range(1, tail + 4):
            ws.cell(rr, col).border = _BORDER

    if report is not None and sheet_name == "合并":
        note = ws.cell(r_course + 2, 1,
                       "注：识别性等级 A=可分辨；B=弱可分辨；C=不可分辨（该目标的数值不应被"
                       "解读为对该目标的独立测量，详见“识别性诊断”工作表）。")
        note.fill = _NOTE_FILL
        ws.merge_cells(start_row=r_course + 2, start_column=1,
                       end_row=r_course + 2, end_column=tail + 3)
        note.alignment = Alignment(wrap_text=True, vertical="center")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 11
    for i in range(len(cols) + 4):
        ws.column_dimensions[get_column_letter(base + i)].width = 11
    ws.freeze_panes = ws.cell(r_pts + 1, 4)


def _write_diagnostics_sheet(ws, bp: Blueprint, result: AttainmentResult,
                             report: DiagnosticReport) -> None:
    ws.cell(1, 1, "达成度识别性诊断").font = Font(bold=True, size=14)
    rows: list[tuple[Any, ...]] = [
        ("指标", "数值", "含义"),
        ("有效秩 (Roy & Vetterli 2007)", round(report.effective_rank, 3),
         f"{report.n_objectives} 个课程目标实际张成的独立维度，奇异值谱熵的指数"),
        ("参与比 (participation ratio)", round(report.participation_ratio, 3),
         "同一谱的另一种刻度，对中间形态的惩罚更重"),
        ("99%方差秩", report.rank99, "解释99%方差所需的维度数"),
        ("目标区分度指数", round(report.separation_index, 4),
         "1 − 目标间相关系数绝对值均值，越接近1越可分辨"),
        ("目标间标准差（学生内）", round(report.between_objective_sd, 5), ""),
        ("学生间标准差", round(report.between_student_sd, 5), ""),
        ("结构性混同目标对", len(report.confounded_pairs),
         "; ".join("/".join(p) for p in report.confounded_pairs) or "无"),
    ]
    r = 3
    for row in rows:
        for c, v in enumerate(row, start=1):
            cell = ws.cell(r, c, v)
            cell.border = _BORDER
            if r == 3:
                cell.font = Font(bold=True)
                cell.fill = _HEAD_FILL
        r += 1

    r += 1
    ws.cell(r, 1, "课程目标达成度与识别性").font = Font(bold=True)
    r += 1
    header = ("课程目标", "指标点", "目标分值", "平均得分", "达成度",
              "95%置信区间", "合格率法", "低于标准概率", "识别性等级")
    for c, v in enumerate(header, start=1):
        cell = ws.cell(r, c, v)
        cell.font = Font(bold=True); cell.fill = _HEAD_FILL; cell.border = _BORDER
    r += 1
    for o in result.objectives:
        grade = report.grades.get(o.id, "?")
        vals = (o.id, o.indicator, round(o.target_points, 3),
                round(o.mean_points, 3), round(o.attainment, 4),
                f"[{o.ci_low:.3f}, {o.ci_high:.3f}]",
                round(o.threshold_attainment, 4), round(o.risk, 4), grade)
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(r, c, v)
            cell.border = _BORDER
            if c == 9:
                cell.fill = _GRADE_FILL.get(grade, _HEAD_FILL)
                cell.font = Font(bold=True)
        r += 1

    r += 1
    ws.cell(r, 1, "目标间相关系数矩阵").font = Font(bold=True)
    r += 1
    for c, oid in enumerate(report.objective_ids, start=2):
        ws.cell(r, c, oid).font = Font(bold=True)
    r += 1
    for i, oid in enumerate(report.objective_ids):
        ws.cell(r, 1, oid).font = Font(bold=True)
        for j in range(len(report.objective_ids)):
            v = report.correlation[i][j]
            cell = ws.cell(r, 2 + j, v)
            cell.border = _BORDER
            if v is not None and i != j and abs(v) >= bp.policy.collinearity_fail:
                cell.fill = _GRADE_FILL["C"]
        r += 1

    r += 1
    ws.cell(r, 1, "诊断明细").font = Font(bold=True)
    r += 1
    for c, v in enumerate(("代码", "级别", "范围", "说明"), start=1):
        cell = ws.cell(r, c, v)
        cell.font = Font(bold=True); cell.fill = _HEAD_FILL; cell.border = _BORDER
    r += 1
    for ck in report.checks:
        for c, v in enumerate((ck.code, ck.level, ck.scope, ck.message), start=1):
            cell = ws.cell(r, c, v)
            cell.border = _BORDER
            cell.alignment = Alignment(wrap_text=(c == 4), vertical="top")
            if c == 2:
                cell.fill = {"fail": _GRADE_FILL["C"], "warn": _GRADE_FILL["B"],
                             "pass": _GRADE_FILL["A"]}[ck.level]
        r += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 90
